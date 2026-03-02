#!/usr/bin/env python3
"""
범용 좌석배치도 → 상품별좌석현황 자동 분류기
색상 기반으로 등급을 자동 감지하여, 어떤 공연장이든 처리 가능.

사용법:
  python3 seat_classifier.py <좌석배치도.xlsx> [시트이름] [관람일] [회차]
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import defaultdict, Counter
import sys
import os
import re


# ─── 색상 유틸 ─────────────────────────────────────────────

def get_color_key(cell):
    """셀의 채우기 색상을 고유 키로 변환"""
    if not cell.fill:
        return None
    fg = cell.fill.fgColor
    if fg is None:
        return None

    # theme 색상 시도
    try:
        if fg.type == 'theme' and fg.theme is not None:
            if cell.fill.fill_type == 'solid':
                return f't{fg.theme}_{fg.tint:.2f}'
            else:
                return f't{fg.theme}_{fg.tint:.2f}'
    except (TypeError, ValueError):
        pass

    # RGB 색상 시도
    try:
        if fg.type == 'rgb' and fg.rgb and isinstance(fg.rgb, str) and fg.rgb != '00000000':
            if cell.fill.fill_type == 'solid':
                return f'rgb_{fg.rgb}'
            return f'rgb_{fg.rgb}'
    except (TypeError, ValueError):
        pass

    return None


# ─── 등급 감지 ─────────────────────────────────────────────

GRADE_KEYWORDS = ['VIP', 'VVIP', 'R', 'S', 'A', 'B', 'C', 'D',
                  'VIP석', 'R석', 'S석', 'A석', 'B석', 'C석',
                  '시야방해R석', '시야방해S석']

GRADE_PRIORITY = {
    'VVIP': 0, 'VVIP석': 0, 'VIP': 1, 'VIP석': 1,
    'R': 2, 'R석': 2, '시야방해R석': 2.5,
    'S': 3, 'S석': 3, '시야방해S석': 3.5,
    'A': 4, 'A석': 4, 'B': 5, 'B석': 5,
    'C': 6, 'C석': 6, 'D': 7, 'D석': 7,
}


BASIC_GRADES = {'VIP', 'VVIP', 'R', 'S', 'A', 'B', 'C', 'D'}


def normalize_grade(name):
    """기본 등급명에 '석' 접미사 추가 (VIP→VIP석, R→R석 등)"""
    if name in BASIC_GRADES:
        return name + '석'
    return name


def detect_legend(ws):
    """시트에서 등급-색상 범례를 자동 감지"""
    legend = {}
    for row in range(1, min(ws.max_row + 1, 20)):
        for col in range(1, min(ws.max_column + 1, 70)):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value).strip() if cell.value else ''
            if val in GRADE_KEYWORDS:
                grade = normalize_grade(val)
                # 1) 셀 자체에 색상이 있는 경우
                ckey = get_color_key(cell)
                if ckey:
                    legend[ckey] = grade
                    continue
                # 2) 인접 셀(좌/우 1~3칸)에 색상이 있는 경우 (부산 스타일)
                for dc in [-1, -2, -3, 1, 2, 3]:
                    adj_col = col + dc
                    if adj_col < 1 or adj_col > ws.max_column:
                        continue
                    adj = ws.cell(row=row, column=adj_col)
                    ckey = get_color_key(adj)
                    if ckey and ckey not in legend:
                        legend[ckey] = grade
                        break
    return legend


# ─── 섹션 감지 (범용) ──────────────────────────────────────

def find_all_labels(ws):
    """시트에서 모든 열/블록 라벨과 층 마커를 찾기"""
    row_labels = []  # (excel_row, excel_col, label_name)
    floor_markers = []  # (excel_row, excel_col, floor_name)

    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 70)):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value).strip() if cell.value else ''

            # 층 마커 (숫자+층 으로 시작하는 셀)
            m = re.match(r'^(\d)층$', val)
            if m:
                floor_markers.append((row, col, val))

            # 열 라벨: "X열" 단독, 또는 "X열 (nnn석)"
            m = re.match(r'^([A-Z])열', val)
            if m and '블록' not in val and '유보' not in val and '알림' not in val:
                row_labels.append((row, col, f'{m.group(1)}열'))

            # 블록 라벨: "X블록(nnn)"
            m = re.search(r'([A-Z])블록\(\d+\)', val)
            if m:
                row_labels.append((row, col, f'{m.group(1)}블록'))

    return row_labels, floor_markers


def _detect_data_islands(ws, data_start, data_end, legend=None):
    """데이터 행에서 등급 색상 셀의 연속 컬럼 그룹(island) 감지"""
    col_has_data = set()
    for r in range(data_start, data_end + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell.value, (int, float)):
                continue
            ckey = get_color_key(cell)
            if not ckey:
                continue
            if legend and ckey not in legend:
                continue
            col_has_data.add(c)

    if not col_has_data:
        return []

    sorted_cols = sorted(col_has_data)
    islands = []
    start = sorted_cols[0]
    prev = start
    for c in sorted_cols[1:]:
        if c - prev > 1:
            islands.append((start, prev))
            start = c
        prev = c
    islands.append((start, prev))
    return islands


def build_sections(ws, row_labels, floor_markers, legend=None):
    """
    라벨들을 섹션(label group + data rows + floor)으로 조직화.
    같은 행에 있는 라벨들은 하나의 섹션 그룹.
    존 경계는 실제 데이터 갭 기반으로 계산.
    """
    # 라벨을 행별로 그룹화
    label_groups = defaultdict(list)
    for r, c, name in row_labels:
        label_groups[r].append((c, name))

    group_rows = sorted(label_groups.keys())

    # ── 인접 라벨 행 병합 (간격 ≤ 2행이면 같은 섹션으로) ──
    merged_groups = []  # list of (list_of_label_rows)
    for gr in group_rows:
        if merged_groups and gr - merged_groups[-1][-1] <= 2:
            merged_groups[-1].append(gr)
        else:
            merged_groups.append([gr])

    sections = []
    for gi, grp in enumerate(merged_groups):
        # 합친 라벨들
        labels = []
        for gr in grp:
            labels.extend(sorted(label_groups[gr], key=lambda x: x[0]))
        labels = sorted(labels, key=lambda x: x[0])

        first_label_row = grp[0]
        last_label_row = grp[-1]

        # 데이터 행 범위: 첫 라벨 행+1 ~ 다음 병합그룹 첫 라벨 행-1
        data_start = first_label_row + 1
        if gi + 1 < len(merged_groups):
            data_end = merged_groups[gi + 1][0] - 1
        else:
            data_end = ws.max_row
        gr = first_label_row

        # 층은 나중에 assign_floors_to_sections에서 결정
        floor = '?'

        is_block = any('블록' in name for _, name in labels)

        if is_block:
            # 블록 스타일: 데이터 island 기반 존 감지 (legend 필터링)
            islands = _detect_data_islands(ws, data_start, data_end, legend)

            if islands and len(islands) >= len(labels):
                # 각 라벨에 가장 가까운 island 할당
                used = set()
                zones = []
                for col, name in labels:
                    best_idx = None
                    best_dist = float('inf')
                    for idx, (isl_s, isl_e) in enumerate(islands):
                        if idx in used:
                            continue
                        # 라벨이 island 내에 있거나 가장 가까운 island
                        if isl_s <= col <= isl_e:
                            dist = 0
                        else:
                            dist = min(abs(col - isl_s), abs(col - isl_e))
                        if dist < best_dist:
                            best_dist = dist
                            best_idx = idx
                    if best_idx is not None:
                        used.add(best_idx)
                        isl_s, isl_e = islands[best_idx]
                        zones.append((isl_s, isl_e, name))
                    else:
                        zones.append((col, col + 10, name))
            else:
                # fallback: 라벨 중간점
                zones = []
                for j, (col, name) in enumerate(labels):
                    zone_start = col
                    if j == len(labels) - 1:
                        zone_end = ws.max_column
                    else:
                        zone_end = labels[j + 1][0] - 2
                    zones.append((zone_start, zone_end, name))
        else:
            # 열 스타일: 데이터 island 기반 존 감지
            islands = _detect_data_islands(ws, data_start, data_end, legend)

            if islands and len(islands) == len(labels):
                zones = [(isl[0], isl[1], labels[j][1]) for j, isl in enumerate(islands)]
            elif islands and len(islands) > len(labels):
                zones = []
                for isl_start, isl_end in islands:
                    isl_center = (isl_start + isl_end) / 2
                    nearest = min(labels, key=lambda l: abs(l[0] - isl_center))
                    zones.append((isl_start, isl_end, nearest[1]))
            else:
                # fallback: 라벨 중간점 기반
                zones = []
                for j, (col, name) in enumerate(labels):
                    if j == 0:
                        zone_start = 1
                    else:
                        prev_col = labels[j - 1][0]
                        zone_start = (prev_col + col) // 2 + 1
                    if j == len(labels) - 1:
                        zone_end = ws.max_column
                    else:
                        next_col = labels[j + 1][0]
                        zone_end = (col + next_col) // 2
                    zones.append((zone_start, zone_end, name))

        sections.append({
            'label_row': gr,
            'labels': labels,
            'zones': zones,
            'data_start': data_start,
            'data_end': data_end,
            'floor': floor,
        })

    return sections


def assign_floors_to_sections(ws, sections, floor_markers):
    """모든 섹션에 층을 배정 (고컬럼 마커 기반)"""
    hi_markers = sorted([(r, c, f) for r, c, f in floor_markers if c > 20], key=lambda x: x[0])

    for section in sections:
        ds = section['data_start']
        de = section['data_end']

        # 1차: 데이터 범위 내 고컬럼 마커 (가장 신뢰)
        for mr, mc, mf in hi_markers:
            if ds <= mr <= de:
                section['floor'] = mf
                break
        else:
            # 2차: 데이터 범위 이후 가장 가까운 고컬럼 마커 (존 기반)
            zone_floor = None
            for mr, mc, mf in hi_markers:
                if mr >= ds:
                    zone_floor = mf
                    break
            if zone_floor:
                section['floor'] = zone_floor
            else:
                section['floor'] = '1층'


# ─── 좌석 파싱 ─────────────────────────────────────────────

def parse_all_seats(ws, legend, sections):
    """모든 섹션에서 좌석을 파싱"""
    seats = []  # (grade, floor, row_label, seat_number)
    seen = set()  # 중복 방지: (excel_row, excel_col)

    for section in sections:
        floor = section['floor']
        zones = section['zones']

        for r in range(section['data_start'], section['data_end'] + 1):
            for c in range(1, ws.max_column + 1):
                if (r, c) in seen:
                    continue

                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, (int, float)):
                    continue

                ckey = get_color_key(cell)
                grade = legend.get(ckey) if ckey else None
                if not grade:
                    continue

                # 컬럼 존 매칭
                row_label = None
                for zone_start, zone_end, name in zones:
                    if zone_start <= c <= zone_end:
                        row_label = name
                        break

                if row_label:
                    seen.add((r, c))
                    seats.append((grade, floor, row_label, int(cell.value)))

    return seats


# ─── 블록 스타일 파싱 (부산 등) ────────────────────────────

def _is_row_col(ws, col, data_start, data_end):
    """열 번호 컬럼인지 확인: 첫 3행이 1,2,3 순차값인지 검사"""
    vals = []
    for r in range(data_start, min(data_start + 3, data_end + 1)):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, (int, float)):
            vals.append(int(v))
        else:
            return False
    return vals == [1, 2, 3] or vals == [1, 2]


def _find_row_col(ws, label_col, section):
    """블록의 열 번호 컬럼을 데이터 기반으로 탐색"""
    ds = section['data_start']
    de = section['data_end']

    # 1차: 헤더에서 '열' 텍스트 찾기
    lr = section['label_row']
    for dc in range(-2, 0):
        adj_col = max(1, label_col + dc)
        adj = ws.cell(row=lr, column=adj_col)
        if adj.value and '열' in str(adj.value):
            if _is_row_col(ws, adj_col, ds, de):
                return adj_col

    # 2차: 주변 컬럼에서 순차값(1,2,3...) 패턴 찾기
    for dc in range(-5, 3):
        check_col = label_col + dc
        if check_col < 1:
            continue
        if _is_row_col(ws, check_col, ds, de):
            return check_col

    # fallback
    return label_col - 1


def parse_block_style(ws, legend, sections):
    """블록 기반 레이아웃에서 좌석 파싱"""
    seats = []
    seen = set()

    # 블록별 열 번호 컬럼 찾기 (데이터 기반 검증)
    block_defs = {}

    for section in sections:
        for col, name in section['labels']:
            if '블록' in name:
                bname = name.replace('블록', '')
                row_col = _find_row_col(ws, col, section)
                block_defs[(section['label_row'], bname)] = {
                    'row_col': row_col,
                    'section': section,
                }

    # 각 블록의 좌석 파싱 (존 기반)
    for section in sections:
        floor = section['floor']
        zones = section['zones']

        for r in range(section['data_start'], section['data_end'] + 1):
            for c in range(1, ws.max_column + 1):
                if (r, c) in seen:
                    continue

                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, (int, float)):
                    continue

                ckey = get_color_key(cell)
                grade = legend.get(ckey) if ckey else None
                if not grade:
                    continue

                # 존 매칭
                zone_name = None
                for zone_start, zone_end, name in zones:
                    if zone_start <= c <= zone_end:
                        zone_name = name
                        break

                if zone_name and '블록' in zone_name:
                    bname = zone_name.replace('블록', '')
                    bdef = block_defs.get((section['label_row'], bname))
                    if bdef:
                        row_cell = ws.cell(row=r, column=bdef['row_col'])
                        if isinstance(row_cell.value, (int, float)):
                            row_num = int(row_cell.value)
                            seen.add((r, c))
                            seats.append((grade, floor, f'{bname}블록{row_num}열', int(cell.value)))

    return seats


# ─── 출력 생성 ─────────────────────────────────────────────

def group_and_sort(seats):
    """좌석 그룹화 & 정렬"""
    groups = defaultdict(set)
    for grade, floor, row_label, seat_num in seats:
        groups[(grade, floor, row_label)].add(seat_num)

    return {k: sorted(v) for k, v in groups.items()}


def sort_key(item):
    grade, floor, row_label = item[0]
    gp = GRADE_PRIORITY.get(grade, 99)
    floor_num = int(re.search(r'\d+', floor).group()) if re.search(r'\d+', floor) else 0

    # 블록+열 정렬
    m = re.match(r'([A-Z])블록(\d+)열', row_label)
    if m:
        return (gp, floor_num, m.group(1), int(m.group(2)))
    # 열 정렬 (알파벳)
    m = re.match(r'([A-Z])열', row_label)
    if m:
        return (gp, floor_num, m.group(1), 0)
    return (gp, floor_num, row_label, 0)


def generate_excel(groups, output_path, event_date='', session=''):
    """상품별좌석현황 엑셀 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '상품별좌석현황'

    headers = ['No', '이용(관람)일', '회차', '좌석등급', '층', '열', '좌석수', '좌석번호']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    grade_fills = {}
    for g in ['VIP', 'VIP석']:
        grade_fills[g] = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    for g in ['R', 'R석', '시야방해R석']:
        grade_fills[g] = PatternFill(start_color='FF0070C0', end_color='FF0070C0', fill_type='solid')
    for g in ['S', 'S석', '시야방해S석']:
        grade_fills[g] = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')
    for g in ['A', 'A석']:
        grade_fills[g] = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    for g in ['B', 'B석']:
        grade_fills[g] = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
    for g in ['C', 'C석']:
        grade_fills[g] = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')

    sorted_items = sorted(groups.items(), key=sort_key)

    row_num = 2
    for idx, (key, seat_list) in enumerate(sorted_items, 1):
        grade, floor, row_label = key
        seat_str = ' '.join(str(s) for s in seat_list)
        values = [idx, event_date, session, grade, floor, row_label, len(seat_list), seat_str]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin
            if col == 4:
                cell.fill = grade_fills.get(grade, PatternFill())
            if col in (1, 3, 5, 7):
                cell.alignment = Alignment(horizontal='center')
        row_num += 1

    ws.cell(row=row_num, column=6, value='합계').font = Font(bold=True)
    ws.cell(row=row_num, column=7, value=f'=SUM(G2:G{row_num - 1})').font = Font(bold=True)

    for letter, w in {'A': 6, 'B': 12, 'C': 14, 'D': 12, 'E': 6, 'F': 14, 'G': 8, 'H': 55}.items():
        ws.column_dimensions[letter].width = w

    wb.save(output_path)
    return row_num - 2


# ─── 메인 ─────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('사용법: python3 seat_classifier.py <좌석배치도.xlsx> [시트이름] [관람일] [회차]')
        sys.exit(1)

    input_file = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    event_date = sys.argv[3] if len(sys.argv) > 3 else ''
    session = sys.argv[4] if len(sys.argv) > 4 else ''

    base = os.path.splitext(os.path.basename(input_file))[0]
    output_dir = os.path.dirname(input_file) or '.'
    output_file = os.path.join(output_dir, f'{base}_좌석현황_자동생성.xlsx')

    print(f'입력: {input_file}')
    print(f'출력: {output_file}')
    print()

    wb = openpyxl.load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f'시트: {ws.title} ({ws.max_row}행 × {ws.max_column}열)')

    # 1. 범례 감지
    legend = detect_legend(ws)
    print(f'\n📋 등급 범례:')
    for ckey, grade in sorted(legend.items(), key=lambda x: GRADE_PRIORITY.get(x[1], 99)):
        print(f'  {grade} ← {ckey}')

    # 2. 구조 감지
    row_labels, floor_markers = find_all_labels(ws)
    is_block_style = any('블록' in name for _, _, name in row_labels)
    print(f'\n🏗️  레이아웃: {"블록 스타일" if is_block_style else "열 스타일"}')
    print(f'  열/블록 라벨: {len(row_labels)}개')
    print(f'  층 마커: {len(floor_markers)}개')

    # 3. 섹션 빌드 & 층 배정
    sections = build_sections(ws, row_labels, floor_markers, legend)
    assign_floors_to_sections(ws, sections, floor_markers)
    print(f'  섹션: {len(sections)}개')
    for s in sections:
        labels = ', '.join(n for _, n in s['labels'])
        print(f'    {s["floor"]} | {labels} | rows {s["data_start"]}~{s["data_end"]}')

    # 4. 좌석 파싱
    if is_block_style:
        seats = parse_block_style(ws, legend, sections)
    else:
        seats = parse_all_seats(ws, legend, sections)

    # 5. 출력
    groups = group_and_sort(seats)
    unique_count = sum(len(v) for v in groups.values())
    print(f'\n🪑 분류된 좌석: {unique_count}석')
    grade_counts = Counter()
    for (g, f, r), slist in groups.items():
        grade_counts[g] += len(slist)
    for g in sorted(grade_counts, key=lambda x: GRADE_PRIORITY.get(x, 99)):
        print(f'  {g}: {grade_counts[g]}석')
    floor_counts = Counter()
    for (g, f, r), slist in groups.items():
        floor_counts[f] += len(slist)
    for f in sorted(floor_counts):
        print(f'  {f}: {floor_counts[f]}석')

    # 6. 출력
    total = generate_excel(groups, output_file, event_date, session)
    print(f'\n✅ {output_file} ({total}행)')


if __name__ == '__main__':
    main()
